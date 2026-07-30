from __future__ import annotations

import tempfile
import unittest
from datetime import date
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook
from PIL import Image

from backend import hydromet_rain_map


class HydrometRainMapTests(unittest.TestCase):
    def test_migrated_assets_cover_every_rain_station(self) -> None:
        stations = hydromet_rain_map.station_names()

        self.assertEqual(21, len(stations))
        self.assertEqual("MataderoSayausi", stations[0])
        self.assertEqual("Irquis", stations[-1])
        self.assertTrue(hydromet_rain_map.BUFFER_PATH.is_file())
        self.assertTrue(hydromet_rain_map.LOGO_PATH.is_file())

    def test_generation_writes_bd_obs_and_the_output_map(self) -> None:
        observations = {
            station: float(index % 8)
            for index, station in enumerate(hydromet_rain_map.station_names())
        }
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            with patch.object(hydromet_rain_map, "REPORT_ROOT", root):
                image_path, workbook_path, preview_path = hydromet_rain_map.generate_rain_map(
                    user_id=12,
                    job_id="test-job",
                    report_date=date(2026, 7, 22),
                    start_time="00:00",
                    end_time="08:00",
                    observations=observations,
                    fetch_basemap=False,
                )

            self.assertEqual(root / "12/jobs/test-job/BD_Obs.xlsx", workbook_path)
            self.assertEqual(root / "12/jobs/test-job/out/mapa_2026-07-22.png", image_path)
            self.assertEqual(
                root / "12/jobs/test-job/out/mapa_limpio_2026-07-22.png",
                preview_path,
            )
            workbook = load_workbook(workbook_path, data_only=True, read_only=True)
            rows = list(workbook.active.iter_rows(values_only=True))
            workbook.close()
            self.assertEqual("Date", rows[0][0])
            self.assertEqual(date(2026, 7, 22), rows[0][1].date())
            self.assertEqual(22, len(rows))
            self.assertEqual(("MataderoSayausi", 0), rows[1])
            self.assertEqual(("Irquis", 4), rows[-1])
            with Image.open(image_path) as image:
                self.assertEqual(hydromet_rain_map.MAP_SIZE, image.size)
                self.assertEqual((2200, 1450), image.size)
                self.assertEqual("RGB", image.mode)
                self.assertEqual((255, 255, 255), image.getpixel((0, 0)))
            with Image.open(preview_path) as preview:
                self.assertEqual(hydromet_rain_map.PLOT_SIZE, preview.size)

    def test_grid_resolution_changes_the_native_interpolation_grid(self) -> None:
        bounds = (-79.35, -3.05, -78.85, -2.70)

        fine = hydromet_rain_map._grid_size(bounds, 0.1)
        coarse = hydromet_rain_map._grid_size(bounds, 0.5)

        self.assertGreater(fine[0], coarse[0])
        self.assertGreater(fine[1], coarse[1])

    def test_idw_power_changes_interpolation_inside_search_radius(self) -> None:
        observations = [
            (hydromet_rain_map.Station("west", -79.2, -2.9, None), 0.0),
            (hydromet_rain_map.Station("east", -79.0, -2.9, None), 30.0),
        ]

        low_power = hydromet_rain_map._cressman_value(
            -79.05,
            -2.9,
            observations,
            search_radius=30,
            p=1,
        )
        high_power = hydromet_rain_map._cressman_value(
            -79.05,
            -2.9,
            observations,
            search_radius=30,
            p=4,
        )

        self.assertGreater(high_power, low_power)

    def test_bounds_are_expanded_without_stretching_the_geography(self) -> None:
        source = (-79.35, -3.05, -78.85, -2.70)
        target_aspect = hydromet_rain_map.PLOT_SIZE[0] / hydromet_rain_map.PLOT_SIZE[1]

        expanded = hydromet_rain_map._expand_bounds_to_aspect(source, target_aspect)

        self.assertAlmostEqual(target_aspect, (expanded[2] - expanded[0]) / (expanded[3] - expanded[1]))
        self.assertLessEqual(expanded[0], source[0])
        self.assertGreaterEqual(expanded[2], source[2])
        self.assertLessEqual(expanded[1], source[1])
        self.assertGreaterEqual(expanded[3], source[3])

    def test_logo_and_design_parameters_control_the_composition(self) -> None:
        observations = dict.fromkeys(hydromet_rain_map.station_names(), 1.0)
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            with (
                patch.object(hydromet_rain_map, "REPORT_ROOT", root),
                patch.object(hydromet_rain_map, "_draw_logo") as draw_logo,
            ):
                image_path, _, _preview_path = hydromet_rain_map.generate_rain_map(
                    user_id=9,
                    job_id="clean-map",
                    report_date=date(2026, 7, 22),
                    start_time="00:00",
                    end_time="08:00",
                    observations=observations,
                    fetch_basemap=False,
                    grid_resolution=1,
                    plot_logo=False,
                    plot_design=False,
                )

            draw_logo.assert_not_called()
            with Image.open(image_path) as image:
                self.assertEqual(hydromet_rain_map.PLOT_SIZE, image.size)


if __name__ == "__main__":
    unittest.main()
