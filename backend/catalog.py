from __future__ import annotations

import threading
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

CATALOG_FILE = Path(__file__).resolve().parent / "data" / "stations.xlsx"
REQUIRED_COLUMNS = {"Código", "Tipo", "X_UTM", "Y_UTM", "Z", "Cuenca"}
_lock = threading.Lock()
_cached_signature: tuple[int, int] | None = None
_cached_catalog: dict[str, dict[str, Any]] = {}


def load_station_catalog() -> dict[str, dict[str, Any]]:
    global _cached_signature, _cached_catalog
    stat = CATALOG_FILE.stat()
    signature = (stat.st_size, stat.st_mtime_ns)
    with _lock:
        if signature == _cached_signature:
            return _cached_catalog
        workbook = load_workbook(CATALOG_FILE, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value).strip() if value is not None else "" for value in next(rows)]
            if not REQUIRED_COLUMNS.issubset(headers):
                missing = ", ".join(sorted(REQUIRED_COLUMNS - set(headers)))
                raise ValueError(f"Faltan columnas en stations.xlsx: {missing}.")
            positions = {header: headers.index(header) for header in REQUIRED_COLUMNS}
            catalog: dict[str, dict[str, Any]] = {}
            for row in rows:
                code_value = row[positions["Código"]]
                if code_value is None or not str(code_value).strip():
                    continue
                code = str(code_value).strip()
                key = normalize_station_code(code)
                if key in catalog:
                    raise ValueError(f"El código {code} está repetido en stations.xlsx.")
                catalog[key] = {
                    "code": code,
                    "type": _text(row[positions["Tipo"]]),
                    "x": _number(row[positions["X_UTM"]]),
                    "y": _number(row[positions["Y_UTM"]]),
                    "z": _number(row[positions["Z"]]),
                    "basin": _text(row[positions["Cuenca"]]),
                }
        finally:
            workbook.close()
        _cached_signature = signature
        _cached_catalog = catalog
        return catalog


def normalize_station_code(value: str) -> str:
    return unicodedata.normalize("NFKC", value).strip().casefold()


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _number(value: Any) -> int | float | str:
    if value is None or value == "":
        return ""
    number = float(value)
    return int(number) if number.is_integer() else number
