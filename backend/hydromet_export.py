from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .desktop_dialogs import choose_save_file


def export_inventory_excel(headers: list[str], rows: list[list[Any]], filename: str) -> dict[str, object]:
    workbook = Workbook(write_only=False)
    sheet = workbook.active
    sheet.title = "Tabla"
    sheet.append(headers)
    for row in rows:
        sheet.append([_safe_excel_value(value) for value in row])

    header_fill = PatternFill("solid", fgColor="405965")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for index, header in enumerate(headers, start=1):
        sample_lengths = [
            len(str(sheet.cell(row=row, column=index).value or ""))
            for row in range(1, min(sheet.max_row, 300) + 1)
        ]
        column_letter = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[column_letter].width = min(
            42,
            max(10, max(sample_lengths, default=len(header)) + 2),
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    stream = BytesIO()
    workbook.save(stream)
    safe_filename = re.sub(r"[^A-Za-z0-9._-]+", "-", filename).strip("-.") or "registro-hidromet"
    output = choose_save_file(
        title="Guardar tabla en Excel",
        suggested_name=f"{safe_filename}.xlsx",
        default_extension=".xlsx",
        file_types=[("Libro de Excel", "*.xlsx")],
    )
    if not output:
        return {"saved": False, "cancelled": True}
    output.write_bytes(stream.getvalue())
    return {"saved": True, "cancelled": False, "filename": output.name, "path": str(output)}


def _safe_excel_value(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value
